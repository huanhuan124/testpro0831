__author__ = 'zenghuan'
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#向excel文件中写入内容
wb = Workbook()
dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
     for col in range(27, 54):
         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)


ws4 = wb.create_sheet(title="Test")
for i in range(1,11):
    ws4.cell(row = i, column =1).value="123"

wb.save(filename = dest_filename)


#读取excel文件中的内容
wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)

for i in range(1,31):
    print(sheet_ranges.cell(column=1,row=i).value)